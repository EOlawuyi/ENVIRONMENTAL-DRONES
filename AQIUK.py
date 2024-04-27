import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import Image
from PIL import Image, ImageTk
from openpyxl import load_workbook
import tkinter.font as tkFont
import numpy as np
import datetime as dt
import math
import openpyxl
import csv
import json
import xlwt



#THIS SOFTWARE BELONGS TO AND IS OWNED BY CHIEF ENGINEER OLOROGUN ENOCH O. EJOFODOMI AND ENGINEER  FRANCIS OLAWUYI.
#THIS SOFTWARE IS OWNED BY DR. JASON ZARA, DR. VESNA ZDERIC, ENGINEER FRANCIS OLAWUYI, DR MICHAEL OLAWUYI, DR. MATTHEW OLAWUYI, DR. ESTHER OLAWUYI, DR. AHMED JENDOUBI, DR. MOHAMED  CHOUIKHA, ENGINEER  ENOCH  EJOFODOMI, EFEJERA EJOFODOMI, AND LUCKY EJOFODOMI.
#THIS SOFTWARE IS AN AIR QUALITY INDEX (AQI) SOFTWARE FOR THE UNITED KINGDOM.
#February 9, 2024.

#create main window of the application
mainwindow = tk.Tk()
mainwindow.geometry("1400x700")
mainwindow.title('AIR QUALITY INDEX (AQI) MAP FOR UK BY OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218')
mainwindow['background']='#7F7FFF'
mainwindow.iconbitmap('companylogo.ico')

global textBox1
global textBox2
global row


def England():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)



    
#    # Load the images of RC14668218 LOGO AND EL ELYON
 #   img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
  #  img2=ImageTk.PhotoImage(file="ELELYON.png")
    # Add the images to the canvas
   # canvas.create_image(50, 50, image=img)
    #canvas.create_image(1320, 50, image=img2)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    # Add the images to the canvas
 #   canvas.create_image(200, 50, image=img3)
  #  canvas.create_image(1235, 50, image=img4)
   # canvas.create_image(680, 50, image=img6)

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

                
    label3a = Label(tk3, text="ENGLAND AQI", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    button1 = Button(tk3, text='CH4',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandMethane)
    button1.place(x=10,y=610)
    button2 = Button(tk3, text='CO',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandCarbonMonoxide)
    button2.place(x=150,y=610)
    button3 = Button(tk3, text='CO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandCarbonDioxide)
    button3.place(x=290,y=610)
    button4 = Button(tk3, text='NH3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandAmmonia)
    button4.place(x=430,y=610)
    button5 = Button(tk3, text='NO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandNitrogenDioxide)
    button5.place(x=570,y=610)
    button6 = Button(tk3, text='O3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandOzone)
    button6.place(x=710,y=610)
    button7 = Button(tk3, text='P.M.2.5',font=('Arial bold', 30), width=6, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandPM25)
    button7.place(x=850,y=610)
    button8 = Button(tk3, text='P.M.10',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandPM10)
    button8.place(x=1015,y=610)
    button9 = Button(tk3, text='SO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=EnglandSulphurDioxide)
    button9.place(x=1155,y=610)



def EnglandMethane():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CH4", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #data1=int(1000) #CH4
    #data2=int(75)  #CO
    #data3=int(130)   #CO2
    #data4=int(187) #NH3
    #data5=int(270)  #NO2
    #data6=int(450)   #O3
    #data7=int(800)    #P.M.2.5
    #data8=int(350)    #P.M.10
    #data9=int(280)    #SO2

    
    #dataframe = openpyxl.load_workbook("AQI.xlsx")
    #dataframe1 = dataframe.active


    #dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    #dataframe1.append(dataB)
    #dataframe.save("AQI.xlsx")

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[0] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[0] > 51) & (ukaqi[0] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[0] > 101) & (ukaqi[0] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[0] > 150) & (ukaqi[0] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[0] > 200) & (ukaqi[0] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[0] > 300) & (ukaqi[0] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[0] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def EnglandCarbonMonoxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[1] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[1] > 51) & (ukaqi[1] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[1] > 101) & (ukaqi[1] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[1] > 150) & (ukaqi[1] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[1] > 200) & (ukaqi[1] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[1] > 300) & (ukaqi[1] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[1] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def EnglandCarbonDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[2] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[2] > 51) & (ukaqi[2] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[2] > 101) & (ukaqi[2] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[2] > 150) & (ukaqi[2] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[2] > 200) & (ukaqi[2] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[2] > 300) & (ukaqi[2] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[2] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def EnglandAmmonia():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NH3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[3] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[3] > 51) & (ukaqi[3] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[3] > 101) & (ukaqi[3] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[3] > 150) & (ukaqi[3] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[3] > 200) & (ukaqi[3] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[3] > 300) & (ukaqi[3] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[3] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def EnglandNitrogenDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[4] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[4] > 51) & (ukaqi[4] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[4] > 101) & (ukaqi[4] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[4] > 150) & (ukaqi[4] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[4] > 200) & (ukaqi[4] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[4] > 300) & (ukaqi[4] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[4] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def EnglandOzone():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="O3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[5] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[5] > 51) & (ukaqi[5] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[5] > 101) & (ukaqi[5] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[5] > 150) & (ukaqi[5] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[5] > 200) & (ukaqi[5] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[5] > 300) & (ukaqi[5] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[5] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def EnglandPM25():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.2.5", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=730,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[6] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[6] > 51) & (ukaqi[6] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[6] > 101) & (ukaqi[6] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[6] > 150) & (ukaqi[6] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[6] > 200) & (ukaqi[6] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[6] > 300) & (ukaqi[6] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[6] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def EnglandPM10():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.10", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[7] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[7] > 51) & (ukaqi[7] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[7] > 101) & (ukaqi[7] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[7] > 150) & (ukaqi[7] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[7] > 200) & (ukaqi[7] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[7] > 300) & (ukaqi[7] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[7] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def EnglandSulphurDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="ENGLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="SO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)

    
    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)


    #Draw Carbon Monoxide AQI Map for the Day
    if(ukaqi[8] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[8] > 51) & (ukaqi[8] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[8] > 101) & (ukaqi[8] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[8] > 150) & (ukaqi[8] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[8] > 200) & (ukaqi[8] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[8] > 300) & (ukaqi[8] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[8] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="ENGLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def Ireland():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)



    
#    # Load the images of RC14668218 LOGO AND EL ELYON
 #   img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
  #  img2=ImageTk.PhotoImage(file="ELELYON.png")
    # Add the images to the canvas
   # canvas.create_image(50, 50, image=img)
    #canvas.create_image(1320, 50, image=img2)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    # Add the images to the canvas
 #   canvas.create_image(200, 50, image=img3)
  #  canvas.create_image(1235, 50, image=img4)
   # canvas.create_image(680, 50, image=img6)

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

                
    label3a = Label(tk3, text="IRELAND AQI", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    button1 = Button(tk3, text='CH4',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandMethane)
    button1.place(x=10,y=610)
    button2 = Button(tk3, text='CO',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandCarbonMonoxide)
    button2.place(x=150,y=610)
    button3 = Button(tk3, text='CO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandCarbonDioxide)
    button3.place(x=290,y=610)
    button4 = Button(tk3, text='NH3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandAmmonia)
    button4.place(x=430,y=610)
    button5 = Button(tk3, text='NO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandNitrogenDioxide)
    button5.place(x=570,y=610)
    button6 = Button(tk3, text='O3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandOzone)
    button6.place(x=710,y=610)
    button7 = Button(tk3, text='P.M.2.5',font=('Arial bold', 30), width=6, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandPM25)
    button7.place(x=850,y=610)
    button8 = Button(tk3, text='P.M.10',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandPM10)
    button8.place(x=1015,y=610)
    button9 = Button(tk3, text='SO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=IrelandSulphurDioxide)
    button9.place(x=1155,y=610)



def IrelandMethane():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CH4", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #data1=int(40) #CH4
    #data2=int(250)  #CO
    #data3=int(380)   #CO2
    #data4=int(460) #NH3
    #data5=int(190)  #NO2
    #data6=int(580)   #O3
    #data7=int(92)    #P.M.2.5
    #data8=int(170)    #P.M.10
    #data9=int(218)    #SO2

    
    #dataframe = openpyxl.load_workbook("AQI.xlsx")
    #dataframe1 = dataframe.active


    #dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    #dataframe1.append(dataB)
    #dataframe.save("AQI.xlsx")

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[9] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[9] > 51) & (ukaqi[9] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[9] > 101) & (ukaqi[9] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[9] > 150) & (ukaqi[9] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[9] > 200) & (ukaqi[9] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[9] > 300) & (ukaqi[9] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[9] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def IrelandCarbonMonoxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[10] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[10] > 51) & (ukaqi[10] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[10] > 101) & (ukaqi[10] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[10] > 150) & (ukaqi[10] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[10] > 200) & (ukaqi[10] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[10] > 300) & (ukaqi[10] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[10] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def IrelandCarbonDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[11] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[11] > 51) & (ukaqi[11] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[11] > 101) & (ukaqi[11] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[11] > 150) & (ukaqi[11] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[11] > 200) & (ukaqi[11] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[11] > 300) & (ukaqi[11] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[11] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def IrelandAmmonia():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NH3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[12] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[12] > 51) & (ukaqi[12] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[12] > 101) & (ukaqi[12] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[12] > 150) & (ukaqi[12] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[12] > 200) & (ukaqi[12] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[12] > 300) & (ukaqi[12] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[12] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def IrelandNitrogenDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[13] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[13] > 51) & (ukaqi[13] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[13] > 101) & (ukaqi[13] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[13] > 150) & (ukaqi[13] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[13] > 200) & (ukaqi[13] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[13] > 300) & (ukaqi[13] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[13] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def IrelandOzone():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="O3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[14] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[14] > 51) & (ukaqi[14] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[14] > 101) & (ukaqi[14] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[14] > 150) & (ukaqi[14] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[14] > 200) & (ukaqi[14] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[14] > 300) & (ukaqi[14] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[14] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def IrelandPM25():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.2.5", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=730,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[15] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[15] > 51) & (ukaqi[15] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[15] > 101) & (ukaqi[15] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[15] > 150) & (ukaqi[15] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[15] > 200) & (ukaqi[15] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[15] > 300) & (ukaqi[15] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[15] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def IrelandPM10():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.10", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[16] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[16] > 51) & (ukaqi[16] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[16] > 101) & (ukaqi[16] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[16] > 150) & (ukaqi[16] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[16] > 200) & (ukaqi[16] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[16] > 300) & (ukaqi[16] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[16] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def IrelandSulphurDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="IRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="SO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[17] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[17] > 51) & (ukaqi[17] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[17] > 101) & (ukaqi[17] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[17] > 150) & (ukaqi[17] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[17] > 200) & (ukaqi[17] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[17] > 300) & (ukaqi[17] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[17] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="IRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def NorthernIreland():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)



    
#    # Load the images of RC14668218 LOGO AND EL ELYON
 #   img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
  #  img2=ImageTk.PhotoImage(file="ELELYON.png")
    # Add the images to the canvas
   # canvas.create_image(50, 50, image=img)
    #canvas.create_image(1320, 50, image=img2)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    # Add the images to the canvas
 #   canvas.create_image(200, 50, image=img3)
  #  canvas.create_image(1235, 50, image=img4)
   # canvas.create_image(680, 50, image=img6)

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

                
    label3a = Label(tk3, text="NORTHERN IRELAND AQI", width=20, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=640,y=110)
    
    button1 = Button(tk3, text='CH4',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandMethane)
    button1.place(x=10,y=610)
    button2 = Button(tk3, text='CO',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandCarbonMonoxide)
    button2.place(x=150,y=610)
    button3 = Button(tk3, text='CO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandCarbonDioxide)
    button3.place(x=290,y=610)
    button4 = Button(tk3, text='NH3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandAmmonia)
    button4.place(x=430,y=610)
    button5 = Button(tk3, text='NO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandNitrogenDioxide)
    button5.place(x=570,y=610)
    button6 = Button(tk3, text='O3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandOzone)
    button6.place(x=710,y=610)
    button7 = Button(tk3, text='P.M.2.5',font=('Arial bold', 30), width=6, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandPM25)
    button7.place(x=850,y=610)
    button8 = Button(tk3, text='P.M.10',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandPM10)
    button8.place(x=1015,y=610)
    button9 = Button(tk3, text='SO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=NorthernIrelandSulphurDioxide)
    button9.place(x=1155,y=610)


def NorthernIrelandMethane():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CH4", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #data1=int(40) #CH4
    #data2=int(250)  #CO
    #data3=int(380)   #CO2
    #data4=int(460) #NH3
    #data5=int(190)  #NO2
    #data6=int(580)   #O3
    #data7=int(92)    #P.M.2.5
    #data8=int(170)    #P.M.10
    #data9=int(218)    #SO2

    
    #dataframe = openpyxl.load_workbook("AQI.xlsx")
    #dataframe1 = dataframe.active


    #dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    #dataframe1.append(dataB)
    #dataframe.save("AQI.xlsx")

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[18] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[18] > 51) & (ukaqi[18] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[18] > 101) & (ukaqi[18] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[18] > 150) & (ukaqi[18] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[18] > 200) & (ukaqi[18] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[18] > 300) & (ukaqi[18] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[18] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def NorthernIrelandCarbonMonoxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[19] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[19] > 51) & (ukaqi[19] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[19] > 101) & (ukaqi[19] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[19] > 150) & (ukaqi[19] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[19] > 200) & (ukaqi[19] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[19] > 300) & (ukaqi[19] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[19] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def NorthernIrelandCarbonDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[20] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[20] > 51) & (ukaqi[20] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[20] > 101) & (ukaqi[20] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[20] > 150) & (ukaqi[20] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[20] > 200) & (ukaqi[20] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[20] > 300) & (ukaqi[20] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[20] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def NorthernIrelandAmmonia():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NH3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[21] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[21] > 51) & (ukaqi[21] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[21] > 101) & (ukaqi[21] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[21] > 150) & (ukaqi[21] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[21] > 200) & (ukaqi[21] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[21] > 300) & (ukaqi[21] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[21] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def NorthernIrelandNitrogenDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[22] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[22] > 51) & (ukaqi[22] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[22] > 101) & (ukaqi[22] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[22] > 150) & (ukaqi[22] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[22] > 200) & (ukaqi[22] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[22] > 300) & (ukaqi[22] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[22] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def NorthernIrelandOzone():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="O3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[23] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[23] > 51) & (ukaqi[23] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[23] > 101) & (ukaqi[23] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[23] > 150) & (ukaqi[23] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[23] > 200) & (ukaqi[23] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[23] > 300) & (ukaqi[23] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[23] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def NorthernIrelandPM25():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.2.5", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=730,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[24] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[24] > 51) & (ukaqi[24] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[24] > 101) & (ukaqi[24] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[24] > 150) & (ukaqi[24] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[24] > 200) & (ukaqi[24] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[24] > 300) & (ukaqi[24] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[24] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def NorthernIrelandPM10():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.10", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[25] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[25] > 51) & (ukaqi[25] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[25] > 101) & (ukaqi[25] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[25] > 150) & (ukaqi[25] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[25] > 200) & (ukaqi[25] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[25] > 300) & (ukaqi[25] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[25] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def NorthernIrelandSulphurDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="NORTHERNIRELAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="SO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)
    

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[26] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[26] > 51) & (ukaqi[26] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[26] > 101) & (ukaqi[26] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[26] > 150) & (ukaqi[26] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[26] > 200) & (ukaqi[26] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[26] > 300) & (ukaqi[26] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[26] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="NORTHERNIRELAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def Scotland():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)



    
#    # Load the images of RC14668218 LOGO AND EL ELYON
 #   img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
  #  img2=ImageTk.PhotoImage(file="ELELYON.png")
    # Add the images to the canvas
   # canvas.create_image(50, 50, image=img)
    #canvas.create_image(1320, 50, image=img2)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    # Add the images to the canvas
 #   canvas.create_image(200, 50, image=img3)
  #  canvas.create_image(1235, 50, image=img4)
   # canvas.create_image(680, 50, image=img6)

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

                
    label3a = Label(tk3, text="SCOTLAND AQI", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    button1 = Button(tk3, text='CH4',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandMethane)
    button1.place(x=10,y=610)
    button2 = Button(tk3, text='CO',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandCarbonMonoxide)
    button2.place(x=150,y=610)
    button3 = Button(tk3, text='CO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandCarbonDioxide)
    button3.place(x=290,y=610)
    button4 = Button(tk3, text='NH3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandAmmonia)
    button4.place(x=430,y=610)
    button5 = Button(tk3, text='NO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandNitrogenDioxide)
    button5.place(x=570,y=610)
    button6 = Button(tk3, text='O3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandOzone)
    button6.place(x=710,y=610)
    button7 = Button(tk3, text='P.M.2.5',font=('Arial bold', 30), width=6, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandPM25)
    button7.place(x=850,y=610)
    button8 = Button(tk3, text='P.M.10',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandPM10)
    button8.place(x=1015,y=610)
    button9 = Button(tk3, text='SO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=ScotlandSulphurDioxide)
    button9.place(x=1155,y=610)



def ScotlandMethane():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CH4", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #data1=int(1000) #CH4
    #data2=int(480)  #CO
    #data3=int(20)   #CO2
    #data4=int(118) #NH3
    #data5=int(240)  #NO2
    #data6=int(301)   #O3
    #data7=int(175)    #P.M.2.5
    #data8=int(160)    #P.M.10
    #data9=int(35)    #SO2

    
    #dataframe = openpyxl.load_workbook("AQI.xlsx")
    #dataframe1 = dataframe.active


    #dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    #dataframe1.append(dataB)
    #dataframe.save("AQI.xlsx")

    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[27] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[27] > 51) & (ukaqi[27] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[27] > 101) & (ukaqi[27] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[27] > 150) & (ukaqi[27] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[27] > 200) & (ukaqi[27] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[27] > 300) & (ukaqi[27] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[27] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def ScotlandCarbonMonoxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[28] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[28] > 51) & (ukaqi[28] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[28] > 101) & (ukaqi[28] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[28] > 150) & (ukaqi[28] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[28] > 200) & (ukaqi[28] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[28] > 300) & (ukaqi[28] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[28] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def ScotlandCarbonDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[29] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[29] > 51) & (ukaqi[29] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[29] > 101) & (ukaqi[29] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[29] > 150) & (ukaqi[29] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[29] > 200) & (ukaqi[29] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[29] > 300) & (ukaqi[29] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[29] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def ScotlandAmmonia():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NH3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[30] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[30] > 51) & (ukaqi[30] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[30] > 101) & (ukaqi[30] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[30] > 150) & (ukaqi[30] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[30] > 200) & (ukaqi[30] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[30] > 300) & (ukaqi[30] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[30] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)







def ScotlandNitrogenDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[31] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[31] > 51) & (ukaqi[31] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[31] > 101) & (ukaqi[31] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[31] > 150) & (ukaqi[31] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[31] > 200) & (ukaqi[31] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[31] > 300) & (ukaqi[31] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[31] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def ScotlandOzone():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="O3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[32] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[32] > 51) & (ukaqi[32] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[32] > 101) & (ukaqi[32] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[32] > 150) & (ukaqi[32] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[32] > 200) & (ukaqi[32] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[32] > 300) & (ukaqi[32] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[32] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)







def ScotlandPM25():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.2.5", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=730,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[33] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[33] > 51) & (ukaqi[33] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[33] > 101) & (ukaqi[33] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[33] > 150) & (ukaqi[33] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[33] > 200) & (ukaqi[33] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[33] > 300) & (ukaqi[33] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[33] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def ScotlandPM10():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.10", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[34] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[34] > 51) & (ukaqi[34] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[34] > 101) & (ukaqi[34] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[34] > 150) & (ukaqi[34] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[34] > 200) & (ukaqi[34] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[34] > 300) & (ukaqi[34] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[34] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def ScotlandSulphurDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="SCOTLAND0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="SO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)



    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[35] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[35] > 51) & (ukaqi[35] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[35] > 101) & (ukaqi[35] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[35] > 150) & (ukaqi[35] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[35] > 200) & (ukaqi[35] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[35] > 300) & (ukaqi[35] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[35] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="SCOTLAND7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def Wales():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)



    
#    # Load the images of RC14668218 LOGO AND EL ELYON
 #   img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
  #  img2=ImageTk.PhotoImage(file="ELELYON.png")
    # Add the images to the canvas
   # canvas.create_image(50, 50, image=img)
    #canvas.create_image(1320, 50, image=img2)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    # Add the images to the canvas
 #   canvas.create_image(200, 50, image=img3)
  #  canvas.create_image(1235, 50, image=img4)
   # canvas.create_image(680, 50, image=img6)

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

                
    label3a = Label(tk3, text="WALES AQI", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    button1 = Button(tk3, text='CH4',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesMethane)
    button1.place(x=10,y=610)
    button2 = Button(tk3, text='CO',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesCarbonMonoxide)
    button2.place(x=150,y=610)
    button3 = Button(tk3, text='CO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesCarbonDioxide)
    button3.place(x=290,y=610)
    button4 = Button(tk3, text='NH3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesAmmonia)
    button4.place(x=430,y=610)
    button5 = Button(tk3, text='NO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesNitrogenDioxide)
    button5.place(x=570,y=610)
    button6 = Button(tk3, text='O3',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesOzone)
    button6.place(x=710,y=610)
    button7 = Button(tk3, text='P.M.2.5',font=('Arial bold', 30), width=6, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesPM25)
    button7.place(x=850,y=610)
    button8 = Button(tk3, text='P.M.10',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesPM10)
    button8.place(x=1015,y=610)
    button9 = Button(tk3, text='SO2',font=('Arial bold', 30), width=5, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=WalesSulphurDioxide)
    button9.place(x=1155,y=610)



def WalesMethane():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CH4", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[36] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[36] > 51) & (ukaqi[36] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[36] > 101) & (ukaqi[36] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[36] > 150) & (ukaqi[36] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[36] > 200) & (ukaqi[36] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[36] > 300) & (ukaqi[36] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[36] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def WalesCarbonMonoxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[37] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[37] > 51) & (ukaqi[37] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[37] > 101) & (ukaqi[37] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[37] > 150) & (ukaqi[37] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[37] > 200) & (ukaqi[37] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[37] > 300) & (ukaqi[37] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[37] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)







def WalesCarbonDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="CO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[38] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[38] > 51) & (ukaqi[38] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[38] > 101) & (ukaqi[38] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[38] > 150) & (ukaqi[38] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[38] > 200) & (ukaqi[38] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[38] > 300) & (ukaqi[38] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[38] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def WalesAmmonia():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NH3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[39] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[39] > 51) & (ukaqi[39] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[39] > 101) & (ukaqi[39] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[39] > 150) & (ukaqi[39] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[39] > 200) & (ukaqi[39] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[39] > 300) & (ukaqi[39] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[39] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





def WalesNitrogenDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="NO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[40] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[40] > 51) & (ukaqi[40] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[40] > 101) & (ukaqi[40] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[40] > 150) & (ukaqi[40] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[40] > 200) & (ukaqi[40] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[40] > 300) & (ukaqi[40] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[40] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)






def WalesOzone():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="O3", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[41] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[41] > 51) & (ukaqi[41] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[41] > 101) & (ukaqi[41] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[41] > 150) & (ukaqi[41] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[41] > 200) & (ukaqi[41] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[41] > 300) & (ukaqi[41] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[41] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)

  




def WalesPM25():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.2.5", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=730,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[42] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[42] > 51) & (ukaqi[42] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[42] > 101) & (ukaqi[42] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[42] > 150) & (ukaqi[42] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[42] > 200) & (ukaqi[42] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[42] > 300) & (ukaqi[42] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[42] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)



def WalesPM10():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="P.M.10", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[43] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[43] > 51) & (ukaqi[43] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[43] > 101) & (ukaqi[43] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[43] > 150) & (ukaqi[43] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[43] > 200) & (ukaqi[43] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[43] > 300) & (ukaqi[43] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[43] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)




def WalesSulphurDioxide():
    #create window
    tk3 = Toplevel()
    tk3.geometry("1400x700")
    tk3.title('AQI FOR UK BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(1400,700)
    tk3.maxsize(1400,700)

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=200)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)


    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=10,y=1)
    img5=PhotoImage(file="ELELYON.png")
    lb5 = Label(tk3, image=img5)
    lb5.img5 = img5 #Save reference to image
    lb5.place(x=1295,y=1)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK
    img3=ImageTk.PhotoImage(file="UKFLAG.png")
    img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
    img5=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")

    #Insert RC14668218 COMPANY HEADING
    label2 = Label(tk3, text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    #CREATE SECOND CANVASS
    canvas2=Canvas(tk3, width=2500, height=700)
    canvas2.place(x=1,y=110)
    #ADD CANVASS BACKGROND
    canvas2.create_image(670, 310, image=img5)


    img6=PhotoImage(file="WHITE BACKGROUND4.png")
    lb6 = Label(tk3, image=img6)
    lb6.img6 = img6 #Save reference to image
    lb6.place(x=1,y=110)

    #UPLOAD ENGLAND IMAGE
    img7=PhotoImage(file="WALES0.png")
    lb7 = Label(tk3, image=img7)
    lb7.img7 = img7 #Save reference to image
    lb7.place(x=1,y=110)

    #UPLOAD COLORCODE IMAGE
    img8=PhotoImage(file="COLORCODE.png")
    lb8 = Label(tk3, image=img8)
    lb8.img8 = img8 #Save reference to image
    lb8.place(x=590,y=180)

    date = dt.datetime.now()
    aqidate = StringVar()
    aqidate = f"{date:%B %d, %Y}"
                
    label3a = Label(tk3, text=aqidate, width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label3a.place(x=760,y=110)
    
    label4a = Label(tk3, text="SO2", width=14, height=1, font=('Arial Bold', 40), bg='#ffffff') 
    label4a.place(x=740,y=570)


    #Obtain AQI Data for UK
    dataframe = openpyxl.load_workbook("AQI.xlsx")
    dataframe1 = dataframe.active

    ukaqi = np.zeros(45)
    print(ukaqi)
    count=0;
    
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            efile = (col[row].value)
            #print(efile)
            if (efile):
                print(efile)
                ukaqi[count] = efile
                count = count+1
    print(ukaqi)



    #Draw Methane AQI Map for the Day
    if(ukaqi[44] < 51):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES1.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[44] > 51) & (ukaqi[44] < 101)):
        #Moderate AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES2.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)        
    if((ukaqi[44] > 101) & (ukaqi[44] < 151)):
        #Unhealthy Sensitive AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES3.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[44] > 150) & (ukaqi[44] < 201)):
        #Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES4.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[44] > 200) & (ukaqi[44] < 301)):
        #Very Unhealthy AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES5.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if((ukaqi[44] > 300) & (ukaqi[44] < 501)):
        #Hazardous AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES6.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)
    if(ukaqi[44] > 501):
        #Good AQI
        #UPLOAD ENGLAND IMAGE
        img10=PhotoImage(file="WALES7.png")
        lb10 = Label(tk3, image=img10)
        lb10.img10 = img10 #Save reference to image
        lb10.place(x=1,y=110)





#Insert RC14668218 COMPANY LOGO AND EL ELYON
# Create a canvas widget
canvas=Canvas(mainwindow, width=1400, height=100)
canvas.pack()

# Load the images of RC14668218 LOGO AND EL ELYON
img=ImageTk.PhotoImage(file="COMPANYLOGO.png")
img2=ImageTk.PhotoImage(file="ELELYON.png")
# Add the images to the canvas
canvas.create_image(50, 50, image=img)
canvas.create_image(1320, 50, image=img2)


#Insert RC14668218 UKFLAG AND UK
# Load the images of RC14668218 UKFLAG AND UK
img3=ImageTk.PhotoImage(file="UKFLAG.png")
img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
img5=ImageTk.PhotoImage(file="AQI BACKGROUND4.png")
img6=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
# Add the images to the canvas
canvas.create_image(200, 50, image=img3)
canvas.create_image(1235, 50, image=img4)
canvas.create_image(680, 50, image=img6)

#Insert RC14668218 COMPANY HEADING
label2 = tk.Label(text="AIR QUALITY INDEX (AQI) MAP FOR UNITED KINGDOM", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 
label2.place(x=260,y=15)
label3 = tk.Label(text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
label3.place(x=257,y=60)

#CREATE SECOND CANVASS
canvas2=Canvas(mainwindow, width=2500, height=700)
canvas2.place(x=1,y=100)
#ADD CANVASS BACKGROND
canvas2.create_image(670, 310, image=img5)

#ADD COMPANY LOGO AND EL ELYON
canvas.create_image(50, 50, image=img)
canvas.create_image(1320, 50, image=img2)

#create UK BUTTONS

button1 = Button(mainwindow, text='ENGLAND',font=('Arial bold', 32), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=England)
button1.place(x=715,y=250)
button2 = Button(mainwindow, text='IRELAND',font=('Arial bold', 32), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=Ireland)
button2.place(x=715,y=340)
button3 = Button(mainwindow, text='NORTHERN IRELAND',font=('Arial bold', 32), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=NorthernIreland)
button3.place(x=715,y=430)
button4 = Button(mainwindow, text='SCOTLAND',font=('Arial bold', 32), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=Scotland)
button4.place(x=715,y=520)
button5 = Button(mainwindow, text='WALES',font=('Arial bold', 32), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=Wales)
button5.place(x=715,y=610)


#create Label and BUTTON for ADD NEW PATIENT 







mainwindow.mainloop()


#button1 = Button(mainwindow, text='PATIENT SEARCH',font=('Arial bold', 17), width=16, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=power_distribution)
